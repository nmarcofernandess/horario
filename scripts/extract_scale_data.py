#!/usr/bin/env python3
"""Extract and normalize escala source files into machine-friendly datasets."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import pdfplumber

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
PROCESSED_DIR = ROOT / "data" / "processed"
DOCS_DIR = ROOT / "docs"

XLSX_MAIN = RAW_DIR / "Horario de trabalho padrao NOVO -2026 - REVISAO DE ESCALA.xlsx"
XLSX_DOM_FOLGAS = RAW_DIR / "DOM E FOLGAS - CAIXA.xlsx"
PDF_SHIFT = RAW_DIR / "escala caixa - rita 1.pdf"
PDF_SUNDAY = RAW_DIR / "escala caixa - rita.pdf"
PRD_SPEC = DOCS_DIR / "PRD_SPEC_Escala_Caixa_Secullum.md"

PT_MONTHS = {
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}

DAY_KEYS = {"SEG", "TER", "QUA", "QUI", "SEX", "SAB", "DOM"}
SUNDAY_PAIR_COLS = [
    ("AP", "AQ", 1),
    ("AR", "AS", 2),
    ("AT", "AU", 3),
    ("AV", "AW", 4),
    ("AX", "AY", 5),
    ("AZ", "BA", 6),
    ("BB", "BC", 7),
    ("BD", "BE", 8),
]

IGNORE_TEXTS = {
    "1A ESCALA",
    "2A ESCALA",
    "3A ESCALA",
    "4A ESCALA",
    "5A ESCALA",
    "6A ESCALA",
    "7A ESCALA",
    "8A ESCALA",
    "TRAB DOM",
    "FOLGA",
}

NAME_ALIASES = {
    "MAYUME": "MAYUMI",
    "MAYUMEI": "MAYUMI",
    "HELO": "HELOISA",
    "ANAJULIA": "ANA JULIA",
}


@dataclass
class SundayRecord:
    source: str
    table_index: int
    scale_index: int
    sunday_raw: str
    sunday_date: str
    employee_raw: str
    employee_norm: str
    folga_raw: str
    folga_date: str


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_text.lower()).strip("_")
    return ascii_text or "sheet"


def normalize_name(value: str) -> str:
    clean = re.sub(r"\s+", " ", value.strip())
    no_accents = (
        unicodedata.normalize("NFKD", clean).encode("ascii", "ignore").decode("ascii")
    )
    canonical = re.sub(r"[^A-Za-z0-9]+", "", no_accents).upper()
    if canonical in NAME_ALIASES:
        return NAME_ALIASES[canonical]
    spaced = re.sub(r"\s+", " ", no_accents).strip().upper()
    return spaced


def looks_like_person_name(value: str) -> bool:
    token = value.strip()
    if not token:
        return False
    if any(ch.isdigit() for ch in token):
        return False
    if parse_pt_date(token, 2026):
        return False
    compact = re.sub(r"[^A-Za-z0-9]+", "", token.upper())
    if compact in IGNORE_TEXTS:
        return False
    return any(ch.isalpha() for ch in token)


def format_timedelta(value: timedelta) -> str:
    total_seconds = int(value.total_seconds())
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)
    hours, rem = divmod(total_seconds, 3600)
    minutes, seconds = divmod(rem, 60)
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"


def parse_hms_minutes(value: str) -> int | None:
    if not value:
        return None
    m = re.match(r"^(-?\d+):([0-5]\d):([0-5]\d)$", value.strip())
    if not m:
        return None
    hours = int(m.group(1))
    minutes = int(m.group(2))
    sign = -1 if hours < 0 else 1
    total = (abs(hours) * 60) + minutes
    return total * sign


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return format_timedelta(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.8f}".rstrip("0").rstrip(".")
    return str(value).strip()


def parse_pt_date(value: str, year: int) -> str:
    raw = value.strip().lower()
    if not raw:
        return ""
    raw = raw.replace(".", "")
    match = re.match(r"^(\d{1,2})/([a-z]{3}|\d{1,2})$", raw)
    if not match:
        return ""
    day_num = int(match.group(1))
    month_token = match.group(2)
    if month_token.isdigit():
        month_num = int(month_token)
    else:
        month_num = PT_MONTHS.get(month_token, 0)
    if not month_num:
        return ""
    try:
        return date(year, month_num, day_num).isoformat()
    except ValueError:
        return ""


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            normalized = {key: as_text(row.get(key, "")) for key in fieldnames}
            writer.writerow(normalized)


def export_workbook(workbook_path: Path, out_dir: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    workbook_slug = slugify(workbook_path.stem)
    workbook_dir = out_dir / "excel_raw" / workbook_slug
    workbook_dir.mkdir(parents=True, exist_ok=True)

    sheet_meta: list[dict[str, Any]] = []
    for sheet in wb.worksheets:
        sheet_slug = slugify(sheet.title)
        csv_path = workbook_dir / f"{sheet_slug}.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            non_empty = 0
            for row in sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                max_col=sheet.max_column,
                values_only=True,
            ):
                normalized_row = [as_text(cell) for cell in row]
                non_empty += sum(1 for cell in normalized_row if cell != "")
                writer.writerow(normalized_row)
        sheet_meta.append(
            {
                "sheet_name": sheet.title,
                "sheet_slug": sheet_slug,
                "rows": sheet.max_row,
                "cols": sheet.max_column,
                "merged_ranges": len(sheet.merged_cells.ranges),
                "non_empty_cells": non_empty,
                "csv_path": str(csv_path.relative_to(ROOT)),
            }
        )

    metadata = {
        "workbook": str(workbook_path.relative_to(ROOT)),
        "export_dir": str(workbook_dir.relative_to(ROOT)),
        "sheet_count": len(wb.worksheets),
        "sheets": sheet_meta,
    }
    meta_path = workbook_dir / "_metadata.json"
    meta_path.write_text(json.dumps(metadata, indent=2, ensure_ascii=False), encoding="utf-8")
    return metadata


def parse_dom_folgas(workbook_path: Path, out_dir: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["Planilha1"]

    header_rows = []
    for row_idx in range(1, ws.max_row + 1):
        b_val = as_text(ws.cell(row_idx, 2).value).upper()
        c_val = as_text(ws.cell(row_idx, 3).value).upper()
        if b_val == "DOM" and c_val == "SEG":
            header_rows.append(row_idx)

    matrix_rows: list[dict[str, Any]] = []
    marker_rows: list[dict[str, Any]] = []

    for idx, header_row in enumerate(header_rows, start=1):
        date_row = header_row + 1
        next_header = header_rows[idx] if idx < len(header_rows) else ws.max_row + 1
        first_col = None
        date_cols: list[int] = []
        for col in range(2, ws.max_column + 1):
            date_value = ws.cell(date_row, col).value
            if isinstance(date_value, (datetime, date)):
                date_cols.append(col)
                if first_col is None:
                    first_col = col
        if not date_cols or first_col is None:
            continue

        employee_rows: list[tuple[int, str]] = []
        for row in range(header_row + 2, next_header):
            name_raw = as_text(ws.cell(row, 1).value)
            if not name_raw:
                continue
            if "SEMANA" in name_raw.upper():
                continue
            employee_rows.append((row, name_raw))

        for col in date_cols:
            date_value = ws.cell(date_row, col).value
            iso_date = date_value.date().isoformat() if isinstance(date_value, datetime) else as_text(date_value)
            day_name = as_text(ws.cell(header_row, col).value).upper()
            week_index = ((col - first_col) // 7) + 1
            for row_idx, employee in employee_rows:
                marker = as_text(ws.cell(row_idx, col).value)
                record = {
                    "block_id": idx,
                    "employee": employee,
                    "employee_norm": normalize_name(employee),
                    "row_index": row_idx,
                    "date": iso_date,
                    "day_name": day_name,
                    "column_index": col,
                    "week_index": week_index,
                    "marker": marker,
                }
                matrix_rows.append(record)
                if marker:
                    marker_rows.append(record)

    matrix_path = out_dir / "dom_folgas_matrix.csv"
    marker_path = out_dir / "dom_folgas_markers.csv"
    write_csv(
        matrix_path,
        matrix_rows,
        [
            "block_id",
            "employee",
            "employee_norm",
            "row_index",
            "date",
            "day_name",
            "column_index",
            "week_index",
            "marker",
        ],
    )
    write_csv(
        marker_path,
        marker_rows,
        [
            "block_id",
            "employee",
            "employee_norm",
            "row_index",
            "date",
            "day_name",
            "column_index",
            "week_index",
            "marker",
        ],
    )

    return {
        "matrix_path": str(matrix_path.relative_to(ROOT)),
        "markers_path": str(marker_path.relative_to(ROOT)),
        "header_blocks": len(header_rows),
        "matrix_rows": len(matrix_rows),
        "marker_rows": len(marker_rows),
        "employees": sorted({row["employee_norm"] for row in matrix_rows}),
    }


def parse_shift_pdf(pdf_path: Path, out_dir: Path) -> dict[str, Any]:
    with pdfplumber.open(pdf_path) as pdf:
        tables = pdf.pages[0].extract_tables() if pdf.pages else []
    if not tables:
        raise RuntimeError(f"No table found in {pdf_path}")

    table = tables[0]
    time_slots = [as_text(cell) for cell in table[0][1:-1]]
    raw_rows = []
    for idx, row in enumerate(table):
        raw_rows.append(
            {
                "row_index": idx,
                **{f"col_{i:02d}": as_text(value) for i, value in enumerate(row, start=1)},
            }
        )

    current_day = ""
    slot_rows: list[dict[str, Any]] = []
    totals_rows: list[dict[str, Any]] = []
    by_day_shift_minutes: dict[tuple[str, str], list[int]] = defaultdict(list)

    for row in table[1:]:
        first_cell = as_text(row[0]).upper()
        total_cell = as_text(row[-1])

        if first_cell[:3] in DAY_KEYS and len(first_cell) >= 3:
            current_day = first_cell[:3]
            continue

        employee = as_text(row[0])
        if not employee:
            continue

        tokens = [as_text(cell).upper() for cell in row[1:-1]]
        non_zero_tokens = [token for token in tokens if token and token != "0"]
        cai_tokens = [token for token in non_zero_tokens if token.startswith("CAI")]
        unique_codes = sorted(set(cai_tokens))
        inferred_shift = unique_codes[0] if len(unique_codes) == 1 else ",".join(unique_codes)
        total_hms = total_cell

        totals_rows.append(
            {
                "day_name": current_day,
                "employee": employee,
                "employee_norm": normalize_name(employee),
                "inferred_shift_code": inferred_shift,
                "day_total_hms": total_hms,
            }
        )

        if inferred_shift and "," not in inferred_shift:
            minutes = parse_hms_minutes(total_hms)
            if minutes is not None:
                by_day_shift_minutes[(current_day, inferred_shift)].append(minutes)

        for slot, token in zip(time_slots, tokens):
            if token and token != "0":
                slot_rows.append(
                    {
                        "day_name": current_day,
                        "employee": employee,
                        "employee_norm": normalize_name(employee),
                        "time_slot": slot,
                        "shift_code": token,
                        "day_total_hms": total_hms,
                    }
                )

    catalog_rows = []
    summary_map: dict[str, list[int]] = defaultdict(list)
    for (day_name, shift_code), minutes_list in sorted(by_day_shift_minutes.items()):
        median = sorted(minutes_list)[len(minutes_list) // 2]
        catalog_rows.append(
            {
                "day_name": day_name,
                "shift_code": shift_code,
                "samples": len(minutes_list),
                "minutes_median": median,
                "hours_median": round(median / 60, 2),
                "minutes_min": min(minutes_list),
                "minutes_max": max(minutes_list),
            }
        )
        summary_map[shift_code].append(median)

    catalog_summary_rows = []
    for shift_code, medians in sorted(summary_map.items()):
        avg_minutes = round(sum(medians) / len(medians), 2)
        catalog_summary_rows.append(
            {
                "shift_code": shift_code,
                "days_covered": len(medians),
                "minutes_avg": avg_minutes,
                "hours_avg": round(avg_minutes / 60, 2),
                "minutes_min": min(medians),
                "minutes_max": max(medians),
            }
        )

    raw_path = out_dir / "pdf_rita1_table_raw.csv"
    slots_path = out_dir / "pdf_rita1_slots.csv"
    totals_path = out_dir / "pdf_rita1_totals.csv"
    catalog_path = out_dir / "pdf_rita1_shift_catalog_by_day.csv"
    catalog_summary_path = out_dir / "pdf_rita1_shift_catalog_summary.csv"
    write_csv(raw_path, raw_rows, list(raw_rows[0].keys()) if raw_rows else ["row_index"])
    write_csv(
        slots_path,
        slot_rows,
        ["day_name", "employee", "employee_norm", "time_slot", "shift_code", "day_total_hms"],
    )
    write_csv(
        totals_path,
        totals_rows,
        ["day_name", "employee", "employee_norm", "inferred_shift_code", "day_total_hms"],
    )
    write_csv(
        catalog_path,
        catalog_rows,
        [
            "day_name",
            "shift_code",
            "samples",
            "minutes_median",
            "hours_median",
            "minutes_min",
            "minutes_max",
        ],
    )
    write_csv(
        catalog_summary_path,
        catalog_summary_rows,
        ["shift_code", "days_covered", "minutes_avg", "hours_avg", "minutes_min", "minutes_max"],
    )

    return {
        "raw_path": str(raw_path.relative_to(ROOT)),
        "slots_path": str(slots_path.relative_to(ROOT)),
        "totals_path": str(totals_path.relative_to(ROOT)),
        "catalog_path": str(catalog_path.relative_to(ROOT)),
        "catalog_summary_path": str(catalog_summary_path.relative_to(ROOT)),
        "slots_rows": len(slot_rows),
        "totals_rows": len(totals_rows),
        "shift_codes": sorted(summary_map.keys()),
    }


def parse_sunday_pdf(pdf_path: Path, out_dir: Path, year: int) -> tuple[dict[str, Any], list[SundayRecord]]:
    with pdfplumber.open(pdf_path) as pdf:
        tables = []
        for page in pdf.pages:
            tables.extend(page.extract_tables() or [])

    records: list[SundayRecord] = []
    for table_index, table in enumerate(tables, start=1):
        if not table or len(table) < 2:
            continue

        first_row = table[0]
        max_pairs = len(first_row) // 2
        sunday_headers = []
        for pair_idx in range(max_pairs):
            sunday_raw = as_text(first_row[pair_idx * 2])
            if not sunday_raw:
                continue
            sunday_date = parse_pt_date(sunday_raw, year)
            if not sunday_date:
                continue
            sunday_headers.append((pair_idx + 1, sunday_raw, sunday_date))

        for row in table[1:]:
            for pair_idx, sunday_raw, sunday_date in sunday_headers:
                col = (pair_idx - 1) * 2
                if col + 1 >= len(row):
                    continue
                employee_raw = as_text(row[col])
                folga_raw = as_text(row[col + 1])
                if not employee_raw:
                    continue
                employee_norm = normalize_name(employee_raw)
                folga_date = parse_pt_date(folga_raw, year)
                records.append(
                    SundayRecord(
                        source="pdf",
                        table_index=table_index,
                        scale_index=pair_idx,
                        sunday_raw=sunday_raw,
                        sunday_date=sunday_date,
                        employee_raw=employee_raw,
                        employee_norm=employee_norm,
                        folga_raw=folga_raw,
                        folga_date=folga_date,
                    )
                )

    csv_path = out_dir / "pdf_rita_sunday_rotation.csv"
    write_csv(
        csv_path,
        [record.__dict__ for record in records],
        [
            "source",
            "table_index",
            "scale_index",
            "sunday_raw",
            "sunday_date",
            "employee_raw",
            "employee_norm",
            "folga_raw",
            "folga_date",
        ],
    )

    sunday_summary = []
    grouped: dict[str, list[SundayRecord]] = defaultdict(list)
    for rec in records:
        grouped[rec.sunday_date].append(rec)
    for sunday_date, items in sorted(grouped.items()):
        sunday_summary.append(
            {
                "sunday_date": sunday_date,
                "headcount": len(items),
                "employees": ", ".join(sorted({r.employee_norm for r in items})),
                "folga_dates": ", ".join(sorted({r.folga_date for r in items if r.folga_date})),
            }
        )
    summary_path = out_dir / "pdf_rita_sunday_rotation_summary.csv"
    write_csv(summary_path, sunday_summary, ["sunday_date", "headcount", "employees", "folga_dates"])

    metadata = {
        "rotation_path": str(csv_path.relative_to(ROOT)),
        "summary_path": str(summary_path.relative_to(ROOT)),
        "rows": len(records),
        "sundays": len(grouped),
        "employees": sorted({r.employee_norm for r in records}),
    }
    return metadata, records


def parse_sunday_from_caixas(
    workbook_path: Path, out_dir: Path
) -> tuple[dict[str, Any], list[SundayRecord]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb["CAIXAS -2026- ESC"]

    header_rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        values = []
        for sunday_col, _, _ in SUNDAY_PAIR_COLS:
            values.append(ws[f"{sunday_col}{row}"].value)
        if values and all(isinstance(v, (datetime, date)) for v in values):
            header_rows.append(row)

    records: list[SundayRecord] = []
    for idx, header_row in enumerate(header_rows):
        next_header = header_rows[idx + 1] if idx + 1 < len(header_rows) else ws.max_row + 1
        sunday_dates: dict[int, str] = {}
        sunday_raw: dict[int, str] = {}
        for sunday_col, _, scale_idx in SUNDAY_PAIR_COLS:
            value = ws[f"{sunday_col}{header_row}"].value
            text_raw = as_text(value)
            sunday_raw[scale_idx] = text_raw
            if isinstance(value, datetime):
                sunday_dates[scale_idx] = value.date().isoformat()
            elif isinstance(value, date):
                sunday_dates[scale_idx] = value.isoformat()
            else:
                sunday_dates[scale_idx] = ""

        for row in range(header_row + 1, next_header):
            for sunday_col, folga_col, scale_idx in SUNDAY_PAIR_COLS:
                employee_raw = as_text(ws[f"{sunday_col}{row}"].value)
                folga_value = ws[f"{folga_col}{row}"].value
                if not employee_raw:
                    continue
                if not looks_like_person_name(employee_raw):
                    continue
                if not isinstance(folga_value, (datetime, date)):
                    continue
                folga_date = (
                    folga_value.date().isoformat()
                    if isinstance(folga_value, datetime)
                    else folga_value.isoformat()
                )
                records.append(
                    SundayRecord(
                        source="xlsx_caixas",
                        table_index=header_row,
                        scale_index=scale_idx,
                        sunday_raw=sunday_raw.get(scale_idx, ""),
                        sunday_date=sunday_dates.get(scale_idx, ""),
                        employee_raw=employee_raw,
                        employee_norm=normalize_name(employee_raw),
                        folga_raw=as_text(folga_value),
                        folga_date=folga_date,
                    )
                )

    deduped = list(
        {
            (
                r.sunday_date,
                r.employee_norm,
                r.folga_date,
                r.scale_index,
            ): r
            for r in records
        }.values()
    )

    csv_path = out_dir / "xlsx_caixas_sunday_rotation.csv"
    write_csv(
        csv_path,
        [record.__dict__ for record in sorted(deduped, key=lambda r: (r.sunday_date, r.scale_index, r.employee_norm))],
        [
            "source",
            "table_index",
            "scale_index",
            "sunday_raw",
            "sunday_date",
            "employee_raw",
            "employee_norm",
            "folga_raw",
            "folga_date",
        ],
    )

    metadata = {
        "rotation_path": str(csv_path.relative_to(ROOT)),
        "rows": len(deduped),
        "header_blocks": len(header_rows),
        "employees": sorted({r.employee_norm for r in deduped}),
    }
    return metadata, deduped


def compare_sunday_sources(
    pdf_records: list[SundayRecord], xlsx_records: list[SundayRecord], out_dir: Path
) -> dict[str, Any]:
    def as_set(records: list[SundayRecord]) -> set[tuple[str, str, str]]:
        return {
            (record.sunday_date, record.employee_norm, record.folga_date)
            for record in records
            if record.sunday_date and record.employee_norm and record.folga_date
        }

    pdf_set = as_set(pdf_records)
    xlsx_set = as_set(xlsx_records)
    only_pdf = sorted(pdf_set - xlsx_set)
    only_xlsx = sorted(xlsx_set - pdf_set)
    intersection = sorted(pdf_set & xlsx_set)

    compare_path = out_dir / "sunday_rotation_source_compare.json"
    payload = {
        "pdf_records": len(pdf_set),
        "xlsx_records": len(xlsx_set),
        "matches": len(intersection),
        "only_pdf": only_pdf,
        "only_xlsx": only_xlsx,
    }
    compare_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    payload["path"] = str(compare_path.relative_to(ROOT))
    return payload


def write_logic_report(
    report_path: Path,
    extract_summary: dict[str, Any],
) -> None:
    dom_meta = extract_summary["dom_folgas"]
    shift_meta = extract_summary["pdf_shift"]
    sunday_pdf_meta = extract_summary["pdf_sunday"]
    sunday_xlsx_meta = extract_summary["xlsx_sunday"]
    compare_meta = extract_summary["sunday_compare"]

    lines = [
        "# Logica Extraida dos Arquivos (Base Streamlit)",
        "",
        "## 1) O que cada arquivo representa",
        f"- `{XLSX_MAIN}`: planilha mestra de escalas por setor; a aba `CAIXAS -2026- ESC` contem os turnos CAI e tambem o quadro de domingo/folga na faixa `AP:BE`.",
        f"- `{XLSX_DOM_FOLGAS}`: matriz compacta de semanas com marcadores numericos (`5` e `6`) por colaborador.",
        f"- `{PDF_SHIFT}`: grade de turnos de segunda a sabado em janelas de 30 minutos (08:00 a 19:30) com totais diarios por colaborador.",
        f"- `{PDF_SUNDAY}`: rodizio de domingos (8 escalas por bloco) + data de folga compensatoria de cada pessoa.",
        f"- `{PRD_SPEC}`: regras de negocio do problema (R1/R2/R3) e alvo funcional para o app.",
        "",
        "## 2) Logica inferida de quem montou a escala",
        "- Segunda a sabado: horarios fixos por codigo (`CAI1`...`CAI6`) sem alteracao estrutural.",
        "- Domingo: escalado em rodizio, com 3 pessoas por domingo na maior parte dos casos.",
        "- Cada pessoa que trabalha no domingo recebe uma data de folga associada (compensacao), registrada no mesmo bloco.",
        "- A planilha `DOM E FOLGAS` parece funcionar como resumo semanal dos marcadores de turno/compensacao por colaborador.",
        "- A aba `CAIXAS -2026- ESC` replica os mesmos dados de rodizio de domingo que aparecem no PDF de domingo.",
        "",
        "## 3) Evidencias extraidas",
        f"- Marcadores `DOM E FOLGAS`: {dom_meta['marker_rows']} marcadores nao vazios em {dom_meta['matrix_rows']} celulas de matriz.",
        f"- Turnos detectados no PDF de grade: {', '.join(shift_meta['shift_codes'])}.",
        f"- Linhas de slots (30 min) extraidas: {shift_meta['slots_rows']}.",
        f"- Rodizio de domingo (PDF): {sunday_pdf_meta['rows']} linhas, {sunday_pdf_meta['sundays']} domingos.",
        f"- Rodizio de domingo (XLSX/CAIXAS): {sunday_xlsx_meta['rows']} linhas deduplicadas.",
        f"- Interseccao PDF x XLSX: {compare_meta['matches']} registros iguais ({compare_meta['pdf_records']} PDF vs {compare_meta['xlsx_records']} XLSX).",
        "",
        "## 4) Artefatos prontos para o Streamlit",
        "- `data/processed/dom_folgas_matrix.csv`: grade completa (inclui vazios).",
        "- `data/processed/dom_folgas_markers.csv`: somente marcadores preenchidos.",
        "- `data/processed/pdf_rita1_slots.csv`: slots de 30 min por colaborador.",
        "- `data/processed/pdf_rita1_totals.csv`: total diario por colaborador.",
        "- `data/processed/pdf_rita1_shift_catalog_by_day.csv`: duracao por codigo/dia.",
        "- `data/processed/pdf_rita_sunday_rotation.csv`: domingo + folga (fonte PDF).",
        "- `data/processed/xlsx_caixas_sunday_rotation.csv`: domingo + folga (fonte XLSX).",
        "- `data/processed/sunday_rotation_source_compare.json`: divergencias entre fontes.",
        "- `data/processed/excel_raw/*`: export bruto de todas as abas em CSV.",
        "",
        "## 5) Pontos que ainda exigem decisao de produto",
        "- O significado exato dos marcadores `5` e `6` deve ser confirmado (turno, dia de folga ou outro codigo operacional).",
        "- Qual corte oficial de apuracao semanal sera usado no app (`MON_SUN` ou `SUN_SAT`).",
        "- Se o app deve apenas validar/corrigir uma escala existente ou gerar escala nova automaticamente.",
    ]
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run(year: int) -> dict[str, Any]:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)

    summary: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "inputs": {
            "xlsx_main": str(XLSX_MAIN.relative_to(ROOT)),
            "xlsx_dom_folgas": str(XLSX_DOM_FOLGAS.relative_to(ROOT)),
            "pdf_shift": str(PDF_SHIFT.relative_to(ROOT)),
            "pdf_sunday": str(PDF_SUNDAY.relative_to(ROOT)),
            "prd_spec": str(PRD_SPEC.relative_to(ROOT)),
        },
    }

    summary["xlsx_main_export"] = export_workbook(XLSX_MAIN, PROCESSED_DIR)
    summary["xlsx_dom_folgas_export"] = export_workbook(XLSX_DOM_FOLGAS, PROCESSED_DIR)
    summary["dom_folgas"] = parse_dom_folgas(XLSX_DOM_FOLGAS, PROCESSED_DIR)
    summary["pdf_shift"] = parse_shift_pdf(PDF_SHIFT, PROCESSED_DIR)
    sunday_pdf_meta, sunday_pdf_records = parse_sunday_pdf(PDF_SUNDAY, PROCESSED_DIR, year)
    summary["pdf_sunday"] = sunday_pdf_meta
    sunday_xlsx_meta, sunday_xlsx_records = parse_sunday_from_caixas(XLSX_MAIN, PROCESSED_DIR)
    summary["xlsx_sunday"] = sunday_xlsx_meta
    summary["sunday_compare"] = compare_sunday_sources(
        sunday_pdf_records, sunday_xlsx_records, PROCESSED_DIR
    )

    report_path = DOCS_DIR / "LOGICA_EXTRAIDA_ESTRUTURA.md"
    write_logic_report(report_path, summary)
    summary["logic_report"] = str(report_path.relative_to(ROOT))

    summary_path = PROCESSED_DIR / "extract_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    summary["summary_path"] = str(summary_path.relative_to(ROOT))
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--year",
        type=int,
        default=2026,
        help="Default year used to parse dd/mmm dates without explicit year.",
    )
    args = parser.parse_args()
    summary = run(year=args.year)
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
